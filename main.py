# ═══════════════════════════════════════════════════════
#  TGSender Backend — FastAPI + Telethon
#  pip install fastapi uvicorn telethon openpyxl aiofiles python-multipart
# ═══════════════════════════════════════════════════════

import asyncio
import os
import random
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import uvicorn
from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from telethon import TelegramClient, errors
import openpyxl

# ── CONFIG ──────────────────────────────────────────────
# Задайте переменные окружения API_ID и API_HASH
# Локально: создайте файл .env или экспортируйте в терминале
# Railway/сервер: добавьте в Variables в панели управления
API_ID   = int(os.environ.get("API_ID", "0"))
API_HASH = os.environ.get("API_HASH", "")
SESSION  = os.environ.get("SESSION_NAME", "tgsender")

app = FastAPI(title="TGSender API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

client: Optional[TelegramClient] = None
mailings: dict = {}
phone_code_hashes = {}


# ═══════════════════════════════════════════════════════
#  MODELS
# ═══════════════════════════════════════════════════════
class PhoneRequest(BaseModel):
    phone: str

class CodeRequest(BaseModel):
    phone: str
    phone_code_hash: str
    code: str

class CreateMailingRequest(BaseModel):
    name: str
    message: str
    file_path: str
    delay_seconds: int = 30
    delay_random: bool = True
    auto_strategy: bool = True


# ═══════════════════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════════════════
@app.post("/api/auth/send-code")
async def send_code(req: PhoneRequest):
    global client
    client = TelegramClient(SESSION, API_ID, API_HASH)
    await client.connect()
    result = await client.send_code_request(req.phone)
    phone_code_hashes[req.phone] = result.phone_code_hash
    return {"ok": True, "phone_code_hash": result.phone_code_hash}


@app.post("/api/auth/verify-code")
async def verify_code(req: CodeRequest):
    try:
        await client.sign_in(req.phone, req.code, phone_code_hash=req.phone_code_hash)
        me = await client.get_me()
        return {"ok": True, "user": {
            "id": me.id, "name": f"{me.first_name} {me.last_name or ''}".strip(),
            "username": me.username or "", "phone": me.phone,
        }}
    except errors.SessionPasswordNeededError:
        raise HTTPException(400, "Нужен пароль 2FA")
    except errors.PhoneCodeInvalidError:
        raise HTTPException(400, "Неверный код")


@app.post("/api/auth/2fa")
async def two_fa(password: str):
    await client.sign_in(password=password)
    me = await client.get_me()
    return {"ok": True, "user": {"id": me.id, "phone": me.phone}}


@app.get("/api/auth/me")
async def get_me():
    if not client or not await client.is_user_authorized():
        raise HTTPException(401, "Не авторизован")
    me = await client.get_me()
    return {"id": me.id, "name": f"{me.first_name} {me.last_name or ''}".strip(),
            "username": me.username or "", "phone": me.phone}


@app.post("/api/auth/logout")
async def logout():
    global client
    if client:
        await client.log_out()
        client = None
    return {"ok": True}


# ═══════════════════════════════════════════════════════
#  FILE UPLOAD
# ═══════════════════════════════════════════════════════
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    content = await file.read()
    path = Path(f"uploads/{file.filename}")
    path.parent.mkdir(exist_ok=True)
    path.write_bytes(content)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Файл пустой")

    headers = [str(h or f"col_{i}").strip().lower() for i, h in enumerate(rows[0])]
    data = [
        {headers[i]: str(v or "").strip() for i, v in enumerate(row)}
        for row in rows[1:] if any(row)
    ]
    return {"filename": file.filename, "total": len(data),
            "columns": headers, "preview": data[:3], "path": str(path)}


# ═══════════════════════════════════════════════════════
#  MAILINGS
# ═══════════════════════════════════════════════════════
@app.post("/api/mailings")
async def create_mailing(req: CreateMailingRequest, background_tasks: BackgroundTasks):
    recipients = []
    if req.file_path and Path(req.file_path).exists():
        wb = openpyxl.load_workbook(req.file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h or "").strip().lower() for h in rows[0]]
        recipients = [
            {headers[i]: str(v or "").strip() for i, v in enumerate(row)}
            for row in rows[1:] if any(row)
        ]

    mailing_id = f"m_{int(time.time())}"
    mailings[mailing_id] = {
        "id": mailing_id, "name": req.name, "message": req.message,
        "status": "running", "total": len(recipients), "sent": 0,
        "delivered": 0, "no_telegram": 0, "errors": 0,
        "recipients": recipients, "delay_seconds": req.delay_seconds,
        "delay_random": req.delay_random, "auto_strategy": req.auto_strategy,
        "created_at": datetime.now().isoformat(), "speed_per_hour": 0, "log": [],
    }
    background_tasks.add_task(run_mailing, mailing_id)
    return {"ok": True, "mailing_id": mailing_id}


@app.get("/api/mailings")
async def list_mailings():
    return [{k: v for k, v in m.items() if k not in ("recipients", "log")} for m in mailings.values()]


@app.get("/api/mailings/{mid}")
async def get_mailing(mid: str):
    m = mailings.get(mid)
    if not m: raise HTTPException(404)
    return {k: v for k, v in m.items() if k != "recipients"}


@app.post("/api/mailings/{mid}/pause")
async def pause_mailing(mid: str):
    m = mailings.get(mid)
    if not m: raise HTTPException(404)
    m["status"] = "paused"
    return {"ok": True}


@app.post("/api/mailings/{mid}/resume")
async def resume_mailing(mid: str, background_tasks: BackgroundTasks):
    m = mailings.get(mid)
    if not m: raise HTTPException(404)
    m["status"] = "running"
    background_tasks.add_task(run_mailing, mid)
    return {"ok": True}


@app.post("/api/mailings/{mid}/stop")
async def stop_mailing(mid: str):
    m = mailings.get(mid)
    if not m: raise HTTPException(404)
    m["status"] = "done"
    return {"ok": True}


@app.delete("/api/mailings/{mid}")
async def delete_mailing(mid: str):
    mailings.pop(mid, None)
    return {"ok": True}


# ═══════════════════════════════════════════════════════
#  CORE SENDER LOGIC — АВТОСТРАТЕГИЯ СКОРОСТИ
# ═══════════════════════════════════════════════════════
async def run_mailing(mailing_id: str):
    """
    Умная рассылка с постепенным набором скорости.

    Схема автостратегии (аналог Wsender):
      День 1:  2–4  сообщений/час  (аккаунт "прогревается")
      День 2:  5–8  сообщений/час
      День 3: 10–15 сообщений/час
      День 4+: 18–25 сообщений/час

    + случайные задержки ±30%
    + редкие длинные паузы (5% шанс, имитация живого человека)
    + FloodWait автообработка (Telegram сам говорит сколько ждать)
    """
    m = mailings.get(mailing_id)
    if not m:
        return

    start_time = time.time()
    sent_this_hour = 0
    hour_start = time.time()
    max_per_hour = 3

    for i, recipient in enumerate(m["recipients"]):
        # Ждём пока снята пауза
        while m["status"] == "paused":
            await asyncio.sleep(3)
        if m["status"] in ("done", "error"):
            break
        if i < m["sent"]:   # уже отправленные (после resume)
            continue

        # ── АВТОСТРАТЕГИЯ ──────────────────────────────
        if m["auto_strategy"]:
            elapsed_days = (time.time() - start_time) / 86400
            if elapsed_days < 1:
                max_per_hour = random.randint(2, 4)
            elif elapsed_days < 2:
                max_per_hour = random.randint(5, 8)
            elif elapsed_days < 3:
                max_per_hour = random.randint(10, 15)
            else:
                max_per_hour = random.randint(18, 25)

            m["speed_per_hour"] = max_per_hour

            # Сброс счётчика раз в час
            if time.time() - hour_start > 3600:
                sent_this_hour = 0
                hour_start = time.time()

            if sent_this_hour >= max_per_hour:
                wait = int(3600 - (time.time() - hour_start)) + random.randint(30, 120)
                m["log"].append(f"⏳ Лимит {max_per_hour}/час. Пауза {wait//60} мин.")
                await asyncio.sleep(wait)
                sent_this_hour = 0
                hour_start = time.time()

        # ── ПОДСТАНОВКА ПЕРЕМЕННЫХ ────────────────────
        text = m["message"]
        for key, val in recipient.items():
            text = text.replace(f"{{{{{key}}}}}", val)

        # ── ОТПРАВКА ─────────────────────────────────
        phone = recipient.get("phone", "").strip()
        try:
            # Ищем пользователя
            try:
                entity = await client.get_entity(phone)
            except Exception:
                m["no_telegram"] += 1
                m["log"].append(f"[{phone}] Нет в Telegram")
                continue

            await client.send_message(entity, text)
            m["sent"] += 1
            m["delivered"] += 1
            sent_this_hour += 1
            m["log"].append(f"[{phone}] ✓ Отправлено ({m['sent']}/{m['total']})")

        except errors.FloodWaitError as e:
            # Telegram требует паузу — обязательно слушаемся!
            wait = e.seconds + random.randint(10, 60)
            m["log"].append(f"⚠️ FloodWait: ждём {wait}с")
            await asyncio.sleep(wait)
            continue  # повторить этот контакт

        except errors.UserPrivacyRestrictedError:
            m["no_telegram"] += 1
            m["log"].append(f"[{phone}] Приватность")

        except errors.PeerFloodError:
            # Серьёзное предупреждение — долгая пауза
            m["log"].append("🚨 PeerFlood! Пауза 2 часа...")
            await asyncio.sleep(7200)

        except Exception as e:
            m["errors"] += 1
            m["log"].append(f"[{phone}] Ошибка: {e}")

        # ── ЗАДЕРЖКА ─────────────────────────────────
        delay = m["delay_seconds"]
        if m["delay_random"]:
            spread = delay * 0.3
            delay = random.uniform(delay - spread, delay + spread)
        if random.random() < 0.05:  # 5% — длинная случайная пауза
            delay += random.randint(120, 300)
        await asyncio.sleep(max(5, delay))

    m["status"] = "done"
    m["log"].append("✅ Рассылка завершена")


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
